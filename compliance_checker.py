#!/usr/bin/env python3
"""
Compliance Checker - Complete HUD ASD Compliance Assessment
Combines HUD ASD calculations with actual distance measurements to determine compliance
"""

import pandas as pd
import numpy as np
import json
import argparse
import sys
from pathlib import Path
from datetime import datetime
import re

# Import distance calculation functionality
# (isolated) sys.path injection removed
from calculate_distances import load_polygon_from_file, calculate_distances_to_polygon


class ComplianceChecker:
    """
    Check compliance by comparing actual distances to HUD ASD required distances
    """
    
    def __init__(self, excel_path: str, hud_results_path: str, polygon_path: str = None):
        """
        Initialize compliance checker
        
        Args:
            excel_path: Path to Excel file with tank data
            hud_results_path: Path to HUD results JSON
            polygon_path: Optional path to polygon coordinates file
        """
        self.excel_path = Path(excel_path)
        self.hud_results_path = Path(hud_results_path)
        self.polygon_path = Path(polygon_path) if polygon_path else None
        
        self.df = None
        self.hud_results = None
        self.polygon_coords = None
        
    def load_data(self):
        """Load all required data"""
        print("📊 Loading data files...")
        
        # Load Excel
        self.df = pd.read_excel(self.excel_path)
        print(f"   ✓ Loaded Excel: {len(self.df)} rows")
        
        # Load HUD results
        with open(self.hud_results_path, 'r') as f:
            self.hud_results = json.load(f)
        print(f"   ✓ Loaded HUD results: {len(self.hud_results)} calculations")
        
        # Load polygon if provided
        if self.polygon_path and self.polygon_path.exists():
            self.polygon_coords = load_polygon_from_file(self.polygon_path)
            print(f"   ✓ Loaded polygon: {len(self.polygon_coords)} vertices")
    
    def calculate_actual_distances(self):
        """Calculate actual distances from polygon to tanks"""
        
        if not self.polygon_coords:
            print("⚠️  No polygon provided, skipping distance calculations")
            return
        
        print("\n📏 Calculating actual distances to polygon boundary...")
        
        # Calculate distances using the catano_distance_calculator
        df_with_distances, stats = calculate_distances_to_polygon(
            self.excel_path,
            self.polygon_coords,
            output_path=None  # Don't save yet
        )
        
        # Update our dataframe with distance results
        self.df = df_with_distances
        
        # Print statistics
        print(f"\n   Distance Statistics:")
        print(f"   • Inside polygon: {stats['inside_polygon']}")
        print(f"   • Outside polygon: {stats['outside_polygon']}")
        if not np.isnan(stats['min_distance']):
            print(f"   • Min distance: {stats['min_distance']:.2f} ft")
            print(f"   • Max distance: {stats['max_distance']:.2f} ft")
            print(f"   • Average: {stats['mean_distance']:.2f} ft")
    
    def add_hud_results(self):
        """Add HUD ASD results to dataframe"""
        
        print("\n🔄 Adding HUD ASD results...")
        
        # Add columns for HUD results
        self.df['ASDPPU (ft)'] = None
        self.df['ASDBPU (ft)'] = None
        self.df['ASDPNPD (ft)'] = None
        self.df['ASDBNPD (ft)'] = None
        self.df['Acceptable Separation Distance Calculated '] = None
        
        # Get site name column
        site_col = self.df.columns[0]
        import re
        def base_name(name: str) -> str:
            s = (name or '').strip()
            s = re.sub(r"\s+[Tt]ank\s+\d+\s*$", "", s)
            return s.strip()
        
        # Match HUD results to Excel rows
        matched_count = 0
        for result in self.hud_results:
            site_name = result.get('name', '')
            site_base = base_name(site_name)
            
            # Find matching row
            mask = (
                self.df[site_col].str.contains(site_name, case=False, na=False) |
                self.df[site_col].str.contains(site_base, case=False, na=False) |
                (self.df[site_col].str.strip().str.lower() == site_base.strip().lower())
            )
            
            if mask.any():
                idx = self.df[mask].index[0]
                
                if 'results' in result and result['results']:
                    asd_values = result['results']
                    
                    # Store individual values for calculations
                    self.df.at[idx, 'ASDPPU (ft)'] = self._parse_distance(asd_values.get('asdppu'))
                    self.df.at[idx, 'ASDBPU (ft)'] = self._parse_distance(asd_values.get('asdbpu'))
                    self.df.at[idx, 'ASDPNPD (ft)'] = self._parse_distance(asd_values.get('asdpnpd'))
                    self.df.at[idx, 'ASDBNPD (ft)'] = self._parse_distance(asd_values.get('asdbnpd'))
                    
                    # Format combined string
                    asd_parts = []
                    if asd_values.get('asdppu'):
                        asd_parts.append(f"ASDPPU - {asd_values.get('asdppu')} ft")
                    if asd_values.get('asdbpu'):
                        asd_parts.append(f"ASDBPU - {asd_values.get('asdbpu')} ft")
                    if asd_values.get('asdpnpd'):
                        asd_parts.append(f"ASDPNPD - {asd_values.get('asdpnpd')} ft")
                    if asd_values.get('asdbnpd'):
                        asd_parts.append(f"ASDBNPD - {asd_values.get('asdbnpd')} ft")
                    
                    self.df.at[idx, 'Acceptable Separation Distance Calculated '] = ' ; '.join(asd_parts)
                    matched_count += 1
        
        print(f"   ✓ Matched {matched_count} HUD results to Excel rows")
    
    def _parse_distance(self, value):
        """Parse distance value to float"""
        if value is None or value == '':
            return None
        
        # Convert to string and extract number
        value_str = str(value).replace('ft', '').replace('feet', '').strip()
        
        try:
            return float(value_str)
        except:
            return None
    
    def determine_compliance(self):
        """
        Determine compliance by comparing actual distance vs required ASD
        Compliance logic:
        - YES: Actual distance > Maximum required ASD
        - NO: Actual distance < Maximum required ASD
        - REVIEW: Unable to determine (missing data)
        - PENDING: Not yet processed
        """
        
        print("\n⚖️ Determining compliance status...")
        
        # Add compliance columns
        self.df['Maximum Required ASD (ft)'] = None
        self.df['Compliance'] = 'PENDING'
        self.df['Compliance Notes'] = ''
        
        # Fill approximate distance from any known distance columns
        if 'Calculated Distance to Polygon (ft)' in self.df.columns:
            self.df['Approximate Distance to Site (appoximately) '] = self.df['Calculated Distance to Polygon (ft)']
        elif 'Distance to Polygon Boundary (ft)' in self.df.columns:
            self.df['Approximate Distance to Site (appoximately) '] = self.df['Distance to Polygon Boundary (ft)']
        
        compliant = 0
        non_compliant = 0
        review = 0
        pending = 0
        
        for idx, row in self.df.iterrows():
            # Get actual distance
            actual_distance = row.get('Approximate Distance to Site (appoximately) ', None)
            
            # Get HUD ASD values
            asdppu = row.get('ASDPPU (ft)', None)
            asdbpu = row.get('ASDBPU (ft)', None)
            asdpnpd = row.get('ASDPNPD (ft)', None)
            asdbnpd = row.get('ASDBNPD (ft)', None)
            
            # Find maximum required distance
            asd_values = [v for v in [asdppu, asdbpu, asdpnpd, asdbnpd] if v is not None]
            
            if asd_values:
                max_asd = max(asd_values)
                self.df.at[idx, 'Maximum Required ASD (ft)'] = max_asd
                
                if actual_distance is not None and not pd.isna(actual_distance):
                    # We have both actual and required distances
                    if actual_distance > max_asd:
                        self.df.at[idx, 'Compliance'] = 'YES'
                        self.df.at[idx, 'Compliance Notes'] = f'Actual ({actual_distance:.1f} ft) > Required ({max_asd:.1f} ft)'
                        compliant += 1
                    else:
                        self.df.at[idx, 'Compliance'] = 'NO'
                        self.df.at[idx, 'Compliance Notes'] = f'Actual ({actual_distance:.1f} ft) < Required ({max_asd:.1f} ft)'
                        non_compliant += 1
                    
                    # Check if inside polygon (additional flag)
                    if row.get('Inside Polygon', False):
                        self.df.at[idx, 'Compliance Notes'] += ' - INSIDE SITE BOUNDARY'
                else:
                    # Have ASD but no actual distance
                    self.df.at[idx, 'Compliance'] = 'REVIEW'
                    self.df.at[idx, 'Compliance Notes'] = 'No actual distance available'
                    review += 1
            else:
                # No ASD calculations
                if pd.notna(row.get('Acceptable Separation Distance Calculated ', None)):
                    self.df.at[idx, 'Compliance'] = 'REVIEW'
                    self.df.at[idx, 'Compliance Notes'] = 'Unable to parse ASD values'
                    review += 1
                else:
                    self.df.at[idx, 'Compliance'] = 'PENDING'
                    self.df.at[idx, 'Compliance Notes'] = 'No HUD calculations available'
                    pending += 1
        
        print(f"\n   Compliance Summary:")
        print(f"   • Compliant (YES): {compliant}")
        print(f"   • Non-Compliant (NO): {non_compliant}")
        print(f"   • Review Needed: {review}")
        print(f"   • Pending: {pending}")
        
        return {
            'compliant': compliant,
            'non_compliant': non_compliant,
            'review': review,
            'pending': pending,
            'total': len(self.df)
        }
    
    def save_compliance_report(self, output_path: str = None):
        """Save the compliance report to Excel"""
        
        if not output_path:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M')
            output_path = self.excel_path.parent / f"Compliance_Report_{timestamp}.xlsx"
        
        print(f"\n💾 Saving compliance report...")
        
        # Reorder columns for better readability
        priority_columns = [
            self.df.columns[0],  # Site name
            'Tank Capacity',
            'Acceptable Separation Distance Calculated ',
            'Maximum Required ASD (ft)',
            'Approximate Distance to Site (appoximately) ',
            'Compliance',
            'Compliance Notes'
        ]
        
        # Get all other columns
        other_columns = [col for col in self.df.columns if col not in priority_columns]
        
        # Reorder dataframe
        ordered_columns = [col for col in priority_columns if col in self.df.columns]
        ordered_columns.extend(other_columns)
        df_ordered = self.df[ordered_columns]
        
        # Save to Excel with formatting
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Main data sheet
            df_ordered.to_excel(writer, sheet_name='Compliance Assessment', index=False)
            
            # Add summary sheet
            self._add_summary_sheet(writer)
            
            # Format compliance column with colors
            self._add_color_coding(writer)
        
        print(f"   ✅ Report saved to: {output_path}")
        return str(output_path)
    
    def _add_summary_sheet(self, writer):
        """Add summary sheet to Excel"""
        
        # Calculate statistics
        compliance_counts = self.df['Compliance'].value_counts()
        
        summary_data = []
        summary_data.append(['HUD ASD Compliance Assessment Report', ''])
        summary_data.append(['', ''])
        summary_data.append(['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M')])
        summary_data.append(['', ''])
        summary_data.append(['Compliance Summary', ''])
        summary_data.append(['Total Sites:', len(self.df)])
        summary_data.append(['Compliant (YES):', compliance_counts.get('YES', 0)])
        summary_data.append(['Non-Compliant (NO):', compliance_counts.get('NO', 0)])
        summary_data.append(['Review Required:', compliance_counts.get('REVIEW', 0)])
        summary_data.append(['Pending Assessment:', compliance_counts.get('PENDING', 0)])
        summary_data.append(['', ''])
        
        # Add distance statistics if available
        if 'Approximate Distance to Site (appoximately) ' in self.df.columns:
            distances = self.df['Approximate Distance to Site (appoximately) '].dropna()
            if len(distances) > 0:
                summary_data.append(['Distance Statistics', ''])
                summary_data.append(['Min Distance:', f'{distances.min():.2f} ft'])
                summary_data.append(['Max Distance:', f'{distances.max():.2f} ft'])
                summary_data.append(['Average Distance:', f'{distances.mean():.2f} ft'])
        
        if 'Maximum Required ASD (ft)' in self.df.columns:
            asds = self.df['Maximum Required ASD (ft)'].dropna()
            if len(asds) > 0:
                summary_data.append(['', ''])
                summary_data.append(['Required ASD Statistics', ''])
                summary_data.append(['Min Required ASD:', f'{asds.min():.2f} ft'])
                summary_data.append(['Max Required ASD:', f'{asds.max():.2f} ft'])
                summary_data.append(['Average Required ASD:', f'{asds.mean():.2f} ft'])
        
        # Critical sites (non-compliant)
        non_compliant_sites = self.df[self.df['Compliance'] == 'NO']
        if len(non_compliant_sites) > 0:
            summary_data.append(['', ''])
            summary_data.append(['⚠️ Non-Compliant Sites', ''])
            for idx, row in non_compliant_sites.head(10).iterrows():
                site_name = row[self.df.columns[0]]
                summary_data.append([f'  • {site_name}', row.get('Compliance Notes', '')])
        
        summary_df = pd.DataFrame(summary_data, columns=['Metric', 'Value'])
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
    
    def _add_color_coding(self, writer):
        """Add color coding to compliance column"""
        
        from openpyxl.styles import PatternFill
        
        worksheet = writer.sheets['Compliance Assessment']
        
        # Find compliance column
        compliance_col_idx = None
        for idx, col_name in enumerate(self.df.columns, 1):
            if col_name == 'Compliance':
                compliance_col_idx = idx
                break
        
        if compliance_col_idx:
            for row_idx, status in enumerate(self.df['Compliance'], 2):
                if pd.notna(status):
                    cell = worksheet.cell(row=row_idx, column=compliance_col_idx)
                    
                    if status == 'YES':
                        # Green for compliant
                        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    elif status == 'NO':
                        # Red for non-compliant
                        cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                    elif status == 'REVIEW':
                        # Orange for review
                        cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                    elif status == 'PENDING':
                        # Yellow for pending
                        cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")


def main():
    """Main execution function"""
    parser = argparse.ArgumentParser(
        description='Check HUD ASD compliance by comparing actual vs required distances',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s tanks.xlsx fast_results.json --polygon catano_boundary.txt
  %(prog)s data.xlsx results.json -o compliance_report.xlsx
  %(prog)s tanks.xlsx results.json --no-distances
        """
    )
    
    parser.add_argument('excel_file', help='Excel file with tank data and coordinates')
    parser.add_argument('hud_results', help='JSON file with HUD ASD results')
    parser.add_argument('--polygon', help='Polygon coordinates file for distance calculation')
    parser.add_argument('-o', '--output', help='Output Excel file name')
    parser.add_argument('--no-distances', action='store_true', 
                       help='Skip distance calculations, use existing distance column')
    
    args = parser.parse_args()
    
    # Validate input files
    if not Path(args.excel_file).exists():
        print(f"❌ Error: Excel file not found: {args.excel_file}")
        sys.exit(1)
    
    if not Path(args.hud_results).exists():
        print(f"❌ Error: HUD results file not found: {args.hud_results}")
        sys.exit(1)
    
    if args.polygon and not Path(args.polygon).exists():
        print(f"❌ Error: Polygon file not found: {args.polygon}")
        sys.exit(1)
    
    print("╔════════════════════════════════════════╗")
    print("║   HUD ASD Compliance Checker           ║")
    print("╚════════════════════════════════════════╝")
    
    # Initialize checker
    checker = ComplianceChecker(
        args.excel_file,
        args.hud_results,
        args.polygon if not args.no_distances else None
    )
    
    try:
        # Load data
        checker.load_data()
        
        # Calculate actual distances if polygon provided
        if args.polygon and not args.no_distances:
            checker.calculate_actual_distances()
        
        # Add HUD results
        checker.add_hud_results()
        
        # Determine compliance
        compliance_stats = checker.determine_compliance()
        
        # Save report
        output_path = checker.save_compliance_report(args.output)
        
        # Print final summary
        print("\n" + "="*50)
        print("COMPLIANCE ASSESSMENT COMPLETE")
        print("="*50)
        
        total = compliance_stats['total']
        compliant = compliance_stats['compliant']
        
        if total > 0:
            compliance_rate = (compliant / total) * 100
            print(f"Overall Compliance Rate: {compliance_rate:.1f}%")
        
        print(f"\nReport saved to: {output_path}")
        print("\n✅ Compliance check complete!")
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
